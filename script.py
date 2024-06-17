from openpyxl import load_workbook
import win32com.client as win32

import os
import re

# Absolute path to the directory containing the spreadsheets from the EUSES corpus
SPREADSHEET_PATH = "[insert path here]"


# Convert .xls file to .xlsx
# Only works on Windows
def convert_xls_to_xlsx(file_name):
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(file_name)

    wb.SaveAs(file_name + "x", FileFormat=51)
    wb.Close()
    excel.Application.Quit()


# Convert all .xls files in a directory to .xlsx
def convert_all_spreadsheets_to_xlsx(dir_path):
    for root, _, files in os.walk(dir_path):
        for file in files:
            file_path = os.path.join(root, file)

            if file_path.endswith(".xls") and not (os.path.isfile(file_path + "x")):
                try:
                    convert_xls_to_xlsx(file_path)
                except:
                    print("Error converting file: " + file_path)


crossSheetNum = 0
rangeRefNum = 0
noneNum = 0
totalNum = 0


# Read a single .xlsx file, and write the cross-sheet and range reference formulas to a file
def read_file(file_name):
    global crossSheetNum, rangeRefNum, noneNum, totalNum

    wb = load_workbook(file_name)
    relative_path = file_name.split(f"{SPREADSHEET_PATH}\\")[1]

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        if "<Worksheet" not in str(ws):
            continue

        localCrossSheetNum = 0
        localRangeRefNum = 0
        for row in ws.iter_rows():
            for cell in row:
                cell_value = cell.value
                cross_sheet_pattern = r"^=[a-zA-Z \d]*![A-Z]\d*"
                range_reference_pattern = r"^=[A-Z]+\([A-Z]+\d+:[A-Z]+\d+\)"
                if not isinstance(cell_value, str):
                    continue

                if re.match(cross_sheet_pattern, cell_value):
                    localCrossSheetNum += 1
                    with open("cross-sheet.txt", "a") as f:
                        f.write(
                            f"{relative_path} Sheet: {sheet}, Cell {cell.coordinate}: {cell_value}\n"
                        )
                elif re.match(range_reference_pattern, cell_value):
                    localRangeRefNum += 1
                    with open("range-ref.txt", "a") as f:
                        f.write(
                            f"{relative_path} Sheet: {sheet}, Cell {cell.coordinate}: {cell_value}\n"
                        )

    if localCrossSheetNum > 0:
        crossSheetNum += 1
    if localRangeRefNum > 0:
        rangeRefNum += 1
    if localCrossSheetNum == 0 and localRangeRefNum == 0:
        noneNum += 1
        with open("none.txt", "a") as f:
            f.write(f"{relative_path}\n")
    totalNum += 1


# Read all .xlsx files in a directory recursively
def read_files_in_directory(dir_path):
    global crossSheetNum, rangeRefNum, noneNum, totalNum
    for root, _, files in os.walk(dir_path):
        for file in files:
            file_path = os.path.join(root, file)

            if file_path.endswith(".xlsx"):
                try:
                    read_file(file_path)
                    print(
                        f"Cross Sheet: {crossSheetNum}, Range Reference: {rangeRefNum}, None: {noneNum}, Total: {totalNum}"
                    )
                except:
                    with open("error.txt", "a") as f:
                        f.write(f"Error reading file: {file_path}\n")


# convert_all_spreadsheets_to_xlsx(SPREADSHEET_PATH)
read_files_in_directory(SPREADSHEET_PATH)
